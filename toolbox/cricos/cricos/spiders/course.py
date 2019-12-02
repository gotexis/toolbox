import scrapy
from scrapy.http import HtmlResponse
from openpyxl import load_workbook, Workbook
import os
from twisted.internet import reactor
from scrapy.crawler import CrawlerRunner
from scrapy.utils.log import configure_logging

mode = 'recrawl_existing'  # recrawl_existing, repeat_everything, skip_existing
search_str = 'https://www.bing.com/search?q='

urllist = []


def query_constuct(search_term1, search_term2, term2_is_site=False):
    if not term2_is_site:
        params = search_term1 + " " + search_term2
    else:
        params = search_term1 + " site:" + search_term2
    return params


# sectoring
for sector in range(1, 10):

    # input_wb = load_workbook('all-' + str(sector - 1) + '.xlsx')
    # output_wb = 'all-' + str(sector) + '.xlsx'
    input_wb = load_workbook('all-25.xlsx')
    output_wb = 'all-25-processed.xlsx'
    ws = input_wb.active
    # rows = ws.iter_rows()

    for i in range(sector * 100, (sector * 100 + 100)):
        # check if already searched
        url1 = ws.cell(column=12, row=i).value
        if url1 != "":
            pass
        else:
            course = ws.cell(column=8, row=i).value
            website = ws.cell(column=11, row=i).value
            institution = ws.cell(column=3, row=i).values
            if not website == "":
                term = query_constuct(course, website, True)
            else:
                term = query_constuct(course, institution, False)

            urllist.append(search_str + term)


class CourseSpider(scrapy.Spider):
    name = "course"

    # def __init__(self, *args, **kwargs):
    #     self.start_urls = kwargs.get('myurls', [])
    #     super(CourseSpider, self).__init__(*args, **kwargs)

    def start_requests(self):
        urls = urllist
        for url in urls:
            yield scrapy.Request(url=url, callback=self.parse)

    def parse(self, response):
        page = response.url.split("=")[-1]
        filename = '%s.xlsx' % page

        # check if school exist
        if response.xpath('//div[@id="pnlErrorMessage"]/text()').extract_first() is not None:
            pass

        else:
            # excel
            wb = Workbook()
            ws = wb.active

            # cricos = response.xpath('//html[1]/body[1]/form[1]/div[3]/div[2]/div[1]/div[1]/div[1]/span[1]').extract()
            # course_lv = response.xpath('/html[1]/body[1]/form[1]/div[1]/div[2]/div[7]/div[1]/table[1]/tr[2]/td[2]/text()').extract()

            headers = [
                'web_id',
                'cricos',
                'name',
                'type',
                'capacity',
                'website',
                'postal',
                'course',
                'level',
                'duration',
                'url_strip'
            ]
            for field in range(0, len(headers)):
                d = ws.cell(row=1, column=field+1, value=headers[field])
                # yield {"field": field}


            # general info
            try:
                url = response.xpath('//a[@id="institutionDetails_hplInstitutionWebAddress"]/text()').extract()[0]
                url_strip = url.strip('https://').strip('http://').strip('www.').strip('/')
            except IndexError:
                url, url_strip = "", ""

            address_list, address = [], ''
            for index, line in enumerate(response.xpath('//span[@id="institutionDetails_lblInstitutionPostalAddress"]/text()')):
                address_line = response.xpath('//span[@id="institutionDetails_lblInstitutionPostalAddress"]/text()')[index].extract()
                address_line = address_line.replace(',', '').strip(' ')
                address_list.append(address_line)
                address = ", ".join(address_list)

            cricos = response.xpath('//span[@id="institutionDetails_lblProviderCode"]/text()').extract()[0]
            capacity = response.xpath('//span[@id="institutionDetails_lblLocationCapacity"]/text()').extract()[0]
            institute_name = response.xpath('//span[@id="institutionDetails_lblInstitutionName"]/text()').extract()[0],
            # if isinstance(institute_name, list):
            #     institute_name = institute_name[0]
            type = response.xpath('//span[@id="institutionDetails_lblInstitutionType"]/text()').extract()[0],
            # end general info

            for row in range(2, (len(response.xpath('//table[@id="courseList_gridSearchResults"]/tr')) + 1)):
                course = [
                    int(page),
                    str(cricos),
                    str(institute_name[0]),  # todo: WTF why this node change
                    str(type[0]),            # todo: WTF why this xpath/node  structure change
                    int(capacity),
                    str(url),
                    str(address),
                    response.xpath('//table[@id="courseList_gridSearchResults"]/tr[' + str(row) + ']/td[1]/text()').extract()[0],
                    response.xpath('//table[@id="courseList_gridSearchResults"]/tr[' + str(row) + ']/td[2]/text()').extract()[0],
                    int(response.xpath('//table[@id="courseList_gridSearchResults"]/tr[' + str(row) + ']/td[3]/text()').extract()[0]),
                    str(url_strip)
                ]

                # excel
                for field in range(0, (len(course))):
                    d = ws.cell(row=row, column=field+1, value=course[field])

            wb.save(output_wb)