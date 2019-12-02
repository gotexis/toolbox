import scrapy
from scrapy.http import HtmlResponse
from openpyxl import load_workbook, Workbook
import os
from twisted.internet import reactor
from scrapy.crawler import CrawlerRunner
from scrapy.utils.log import configure_logging


mode = 'recrawl_existing'  # recrawl_existing, repeat_everything, skip_existing
search_str = 'http://cricos.education.gov.au/Institution/InstitutionDetailsOnePage.aspx?ProviderID='

urllist = []

# mode selection
for item in range(3471, 4000):
    if mode == 'recrawl_existing':
        if os.path.isfile('./output/' + str(item) + '.xlsx'):
            urllist.append(search_str + str(item))
    elif mode == 'repeat_everything':
        urllist.append(search_str + str(item))
    elif mode == 'skip_existing':
        if not os.path.isfile('./output/' + str(item) + '.xlsx'):
            urllist.append(search_str + str(item))
    else:
        pass


class CricosSpider(scrapy.Spider):
    name = "cricos"

    # def __init__(self, *args, **kwargs):
    #     self.start_urls = kwargs.get('myurls', [])
    #     super(CricosSpider, self).__init__(*args, **kwargs)

    def start_requests(self):
        urls = urllist
        for url in urls:
            yield scrapy.Request(url=url, callback=self.parse)

    def parse(self, response):
        page = response.url.split("=")[-1]
        filename = '%s.xlsx' % page

        # check if school exist
        if response.xpath('//div[@id="pnlErrorMessage"]/text()').extract_first() is not None:
            pass

        else:
            # excel
            wb = Workbook()
            ws = wb.active

            # cricos = response.xpath('//html[1]/body[1]/form[1]/div[3]/div[2]/div[1]/div[1]/div[1]/span[1]').extract()
            # course_lv = response.xpath('/html[1]/body[1]/form[1]/div[1]/div[2]/div[7]/div[1]/table[1]/tr[2]/td[2]/text()').extract()

            headers = [
                'web_id',
                'cricos',
                'name',
                'type',
                'capacity',
                'website',
                'postal',
                'course',
                'level',
                'duration',
                'url_strip'
            ]
            for field in range(0, len(headers)):
                d = ws.cell(row=1, column=field+1, value=headers[field])
                # yield {"field": field}


            # general info
            try:
                url = response.xpath('//a[@id="institutionDetails_hplInstitutionWebAddress"]/text()').extract()[0]
                url_strip = url.replace('https://', "").replace('http://', "").replace('www.', '').strip('/')

            except IndexError:
                url, url_strip = "", ""

            address_list, address = [], ''
            for index, line in enumerate(response.xpath('//span[@id="institutionDetails_lblInstitutionPostalAddress"]/text()')):
                address_line = response.xpath('//span[@id="institutionDetails_lblInstitutionPostalAddress"]/text()')[index].extract()
                address_line = address_line.replace(',', '').strip()
                address_list.append(address_line)
                address = ", ".join(address_list)

            cricos = response.xpath('//span[@id="institutionDetails_lblProviderCode"]/text()').extract()[0]
            capacity = response.xpath('//span[@id="institutionDetails_lblLocationCapacity"]/text()').extract()[0]
            institute_name = response.xpath('//span[@id="institutionDetails_lblInstitutionName"]/text()').extract()[0],
            # if isinstance(institute_name, list):
            #     institute_name = institute_name[0]
            type = response.xpath('//span[@id="institutionDetails_lblInstitutionType"]/text()').extract()[0],
            # end general info

            for row in range(2, (len(response.xpath('//table[@id="courseList_gridSearchResults"]/tr')) + 1)):
                course = [
                    int(page),
                    str(cricos),
                    str(institute_name[0]),  # todo: WTF why this node change
                    str(type[0]),            # todo: WTF why this xpath/node  structure change
                    int(capacity),
                    str(url),
                    str(address),
                    response.xpath('//table[@id="courseList_gridSearchResults"]/tr[' + str(row) + ']/td[1]/text()').extract()[0],
                    response.xpath('//table[@id="courseList_gridSearchResults"]/tr[' + str(row) + ']/td[2]/text()').extract()[0],
                    int(response.xpath('//table[@id="courseList_gridSearchResults"]/tr[' + str(row) + ']/td[3]/text()').extract()[0]),
                    str(url_strip)
                ]

                # excel
                for field in range(0, (len(course))):
                    d = ws.cell(row=row, column=field+1, value=course[field])

            wb.save(os.path.join('output/', filename))

        # with open(filename, 'w') as f:
        #     f.write("\n".join(info))

        # self.log('Saved file %s' % filename)
