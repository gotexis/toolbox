import requests
import json
from openpyxl import Workbook, load_workbook
import os
import logging

# logging
logging.basicConfig(filename='example.log', level=logging.DEBUG)

# settings
search_url = "https://api.cognitive.microsoft.com/bing/v7.0/search"
subscription_key1 = "2c71dcdb903c40a0b805679d06d7104f"
subscription_key2 = "d855fdfcae964c34ba1e36b9ad3507e8"
headers = {"Ocp-Apim-Subscription-Key": subscription_key2}


# make search params
def query(search_term1, search_term2, site=False):
    if not site:
        params = {"q": search_term1 + " " + search_term2, "textDecorations": True, "textFormat": "HTML"}
    else:
        params = {"q": search_term1 + " site:" + search_term2, "textDecorations": True, "textFormat": "HTML"}
    return params


# conduct search
def search(search_term1, search_term2, site=False):
    if not site:
        params = {"q": search_term1 + " " + search_term2, "textDecorations": True, "textFormat": "HTML"}
    else:
        params = {"q": search_term1 + " site:" + search_term2, "textDecorations": True, "textFormat": "HTML"}
    response = requests.get(search_url, headers=headers, params=params)
    response.raise_for_status()
    search_results = response.json()
    return search_results


# sectoring
for sector in range(26, 226):

    input_wb = load_workbook('all-' + str(sector - 1) + '.xlsx')
    output_wb = 'all-' + str(sector) + '.xlsx'
    ws = input_wb.active
    # rows = ws.iter_rows()

    for i in range(sector * 100, (sector * 100 + 100)):
        course = ws.cell(column=8, row=i).value
        url = ws.cell(column=11, row=i).value
        institution = ws.cell(column=3, row=i).value
        query = "site:" + url + " " + course
        if not url == "":
            data = search(course, url, True)
        else:
            data = search(course, institution, False)
        # response

        # process response json
        def record_query():
            url1 = data["webPages"]["value"][0]["url"]
            d = ws.cell(column=12, row=i, value=url1)
            url2 = data["webPages"]["value"][1]["url"]
            d = ws.cell(column=13, row=i, value=url2)
            title = data["webPages"]["value"][0]["name"]
            title = title.replace("<b>", "").replace("</b>", "").strip(" ")
            d = ws.cell(column=14, row=i, value=title)
        try:
            record_query()
        except (KeyError, IndexError) as e:
            try:
                data = search(course, institution, False)
                record_query()
            except (KeyError, IndexError) as e:
                # logging.exception('KeyError for ' + str(i))
                with open(str(i) + '.json', 'w') as f:
                    f.write(json.dumps(data, indent=4))
                print(str(e) + " " + str(i) + " row! Saved Json.")
                print(e.args)
        # success: line
        print("Processed #" + str(i) + " line.")
    # success: sector
    input_wb.save(output_wb)
    print("Processed #" + str(sector) + " sector.")


#
# # read json
# data = json.load(open('outputfile.json'))

# shutdown code
# os.system("shutdown /s /t 1")