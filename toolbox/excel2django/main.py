from openpyxl import load_workbook
import os

# HTML append to start/end
start = """{% extends "base.html" %}
{% block content %}

"""

end = """

{% endblock content %}"""

# load workbook
wb = load_workbook(filename = 'input.xlsx', data_only = True)
ws = wb.worksheets[0]
max_row = ws.max_row

views_list = []
url_list = []

# loop through sheet
for x in range(1, max_row + 1):

    # read excel
    template_name = ws.cell(x, 1).value
    HTMLcode = ws.cell(x, 2).value
    HTMLcode_django_tagged = start + HTMLcode + end

    # parse path
    if "/" in template_name:
        path = template_name.split("/")[0]
    else:
        path = ""
    file = template_name.split("/")[-1]

    # path slash or underscore file
    if path != "":
        path_slash_file = path + "/" + file
        path_file = path + "_" + file
    else:
        path_slash_file = file
        path_file = file

    # mkdir and output file
    if path != "":
        if not os.path.exists("output/" + path):
            os.mkdir("output/" + path)
    with open("output/" + path_slash_file + ".html", "w+", encoding='utf-8') as text_file:
        print(HTMLcode_django_tagged, file=text_file)

    # generate views.py and url.py list
    view_entry = "def " + path_file + """(request):
    return render(request, '""" + path_slash_file + ".html')"
    url_entry = "url(r'^" + path_slash_file + "/$', views." + path_file + ", name='" + path_file + "'),"

    views_list.append(view_entry)
    url_list.append(url_entry)

#  output final view_list and url_list
with open("output/" + "view_list.txt", "w+", encoding='utf-8') as text_file:
    print("\n\n\n".join(views_list), file=text_file)

with open("output/" + "url_list.txt", "w+", encoding='utf-8') as text_file:
    print("\n".join(url_list), file=text_file)