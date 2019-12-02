from openpyxl import load_workbook
import os

# load html
with open("sidebar.html", encoding='utf-8') as inputfile:
    html = inputfile.read()


# load workbook
wb = load_workbook(filename = 'URL-template list.xlsx', data_only = True)
ws = wb.worksheets[0]
max_row = ws.max_row


# loop through sheet
for x in range(1, max_row + 1):
    url_old = ws.cell(x, 1).value
    url_new = ws.cell(x, 2).value
    print(url_old)
    print(url_new)
    html = html.replace(url_old, url_new)

with open("output/" + "sidebar_processed.html", "w+", encoding='utf-8') as text_file:
    print(html, file=text_file)
