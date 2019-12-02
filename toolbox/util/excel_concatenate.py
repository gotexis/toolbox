import os
from openpyxl import Workbook, load_workbook
# import openpyxl



file_list = []
for i in range(1, 4000):
    filename = str(i) + ".xlsx"
    excel_file = os.path.join("../cricos/cricos/output/", filename)
    if not os.path.isfile(excel_file):
        pass
    else:
        file_list.append(excel_file)


new_wb = Workbook()
new_ws = new_wb.active

for file in file_list:
    wb = load_workbook(file)
    ws = wb.active
    rows = ws.iter_rows()
    for index, row in enumerate(rows):
        if index > 1:
            item = []
            for cell in row:
                item.append(cell.value)
            new_ws.append(item)

new_wb.save("all.xlsx")
